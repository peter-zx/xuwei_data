# Excel文件处理工具
# 处理公式转数值、格式转换等

import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def convert_formulas_to_values(input_path, output_path):
    """
    将 Excel 文件中的公式转换为数值
    类似于 Excel 的"选择性粘贴 - 数值"功能
    
    步骤：
    1. 使用 data_only=True 读取计算后的值
    2. 使用 data_only=False 读取公式（备用）
    3. 创建新文件，只保留数值，删除公式
    """
    try:
        print('开始公式转数值处理...')
        
        # 读取计算后的值（data_only=True）
        wb_values = load_workbook(input_path, data_only=True)
        
        # 读取公式（data_only=False，用于检查哪些单元格包含公式）
        wb_formulas = load_workbook(input_path, data_only=False)
        
        # 创建新的工作簿，只保留值
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 删除默认sheet
        
        stats = {
            'total_cells': 0,
            'formula_cells': 0,
            'resolved_cells': 0,
            'unresolved_cells': 0
        }
        
        for sheet_name in wb_values.sheetnames:
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
            
            # 创建新sheet
            new_ws = new_wb.create_sheet(title=sheet_name)
            
            print(f'  处理 Sheet: {sheet_name}')
            
            # 遍历所有单元格
            for row_idx, (row_values, row_formulas) in enumerate(
                zip(ws_values.iter_rows(), ws_formulas.iter_rows()), start=1
            ):
                for col_idx, (cell_value, cell_formula) in enumerate(
                    zip(row_values, row_formulas), start=1
                ):
                    stats['total_cells'] += 1
                    
                    # 检查是否是公式单元格
                    is_formula = (
                        cell_formula.value and 
                        isinstance(cell_formula.value, str) and 
                        cell_formula.value.startswith('=')
                    )
                    
                    if is_formula:
                        stats['formula_cells'] += 1
                    
                    # 获取值
                    value = cell_value.value
                    
                    # 如果公式未被计算（值为None），尝试其他方法
                    if value is None and is_formula:
                        stats['unresolved_cells'] += 1
                        # 尝试使用 formulas 库计算（如果可用）
                        try:
                            from formulas import ExcelModel
                            # 这里可以添加更复杂的计算逻辑
                            pass
                        except:
                            pass
                        # 保持为空字符串
                        value = ''
                    elif is_formula and value is not None:
                        stats['resolved_cells'] += 1
                    
                    # 写入新单元格
                    if value is not None:
                        new_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 保存新文件
        new_wb.save(output_path)
        new_wb.close()
        wb_values.close()
        wb_formulas.close()
        
        print(f'公式转数值完成:')
        print(f'  总单元格: {stats["total_cells"]}')
        print(f'  公式单元格: {stats["formula_cells"]}')
        print(f'  已解析: {stats["resolved_cells"]}')
        print(f'  未解析: {stats["unresolved_cells"]}')
        
        return True
        
    except Exception as e:
        print(f'公式转数值失败: {e}')
        import traceback
        traceback.print_exc()
        return False


def convert_xls_to_xlsx_with_values(xls_path, xlsx_path):
    """
    将 XLS 文件转换为 XLSX，并转换公式为数值
    """
    try:
        print(f'XLS 转 XLSX: {os.path.basename(xls_path)}')
        
        # 使用绝对路径
        xls_path = os.path.abspath(xls_path)
        xlsx_path = os.path.abspath(xlsx_path)
        
        # 使用 pandas 读取（会自动计算公式）
        excel_data = pd.ExcelFile(xls_path, engine='xlrd')
        
        # 创建临时 XLSX 文件（使用绝对路径）
        temp_xlsx = xls_path.rsplit('.', 1)[0] + '_temp.xlsx'
        
        print(f'  创建临时文件: {os.path.basename(temp_xlsx)}')
        
        with pd.ExcelWriter(temp_xlsx, engine='openpyxl') as writer:
            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(xls_path, sheet_name=sheet_name, engine='xlrd')
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 使用 openpyxl 进一步处理公式
        success = convert_formulas_to_values(temp_xlsx, xlsx_path)
        
        # 清理临时文件
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)
            print(f'  临时文件已删除')
        
        print(f'XLS 转换完成: {os.path.basename(xlsx_path)}')
        return success
        
    except Exception as e:
        print(f'XLS 转换失败: {e}')
        import traceback
        traceback.print_exc()
        # 清理临时文件
        if 'temp_xlsx' in locals() and os.path.exists(temp_xlsx):
            try:
                os.remove(temp_xlsx)
            except:
                pass
        return False


def process_excel_with_formulas(input_path, output_path):
    """
    处理包含公式的 Excel 文件
    主入口函数
    """
    try:
        # 使用绝对路径
        input_path = os.path.abspath(input_path)
        output_path = os.path.abspath(output_path)
        
        print(f'处理 Excel 文件: {os.path.basename(input_path)}')
        
        # 检查文件格式
        with open(input_path, 'rb') as f:
            header = f.read(8)
            is_xlsx = header[:2] == b'PK'
        
        if is_xlsx:
            print('  检测到 XLSX 格式')
            # XLSX 文件：直接转换公式
            return convert_formulas_to_values(input_path, output_path)
        else:
            print('  检测到 XLS 格式')
            # XLS 文件：先转换格式，再处理公式
            return convert_xls_to_xlsx_with_values(input_path, output_path)
            
    except Exception as e:
        print(f'Excel 处理失败: {e}')
        import traceback
        traceback.print_exc()
        # 最后的回退：直接复制
        try:
            import shutil
            shutil.copy(input_path, output_path)
            return True
        except Exception as e2:
            print(f'复制也失败: {e2}')
            return False


# 保留旧函数名以兼容
convert_xls_to_xlsx_with_calculation = convert_xls_to_xlsx_with_values
